# -*- coding: utf-8 -*-
"""塾の月末請求一覧表を生成するスクリプト。

openpyxl を使ってダミー生徒180名分の請求データを作成し、
「一覧表.xlsx」として保存する。
"""

import datetime
import json
import random
import re
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.properties import Outline, PageSetupProperties
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

random.seed(42)

# ------------------------------------------------------------
# ダミーデータ生成
# ------------------------------------------------------------

SEI_LIST = [
    "佐藤", "鈴木", "高橋", "田中", "伊藤", "渡辺", "山本", "中村", "小林", "加藤",
    "吉田", "山田", "佐々木", "山口", "松本", "井上", "木村", "林", "斎藤", "清水",
    "山崎", "森", "池田", "橋本", "阿部", "石川", "山下", "中島", "石井", "小川",
    "前田", "岡田", "長谷川", "藤田", "後藤", "近藤", "村上", "遠藤", "青木", "坂本",
    "福田", "太田", "西村", "藤井", "岡本", "中野", "小野", "田村", "竹内", "金子",
]

MEI_LIST = [
    "翔太", "陽菜", "大輝", "美咲", "健太", "愛", "拓也", "結衣", "颯太", "さくら",
    "蓮", "陽菜乃", "悠斗", "葵", "大和", "美月", "湊", "花音", "樹", "莉子",
    "陸", "楓", "颯", "凛", "駿", "詩織", "航", "彩花", "翼", "優奈",
    "尊", "咲良", "結翔", "心春", "陽向", "美結", "朝陽", "琴音", "悠真", "千尋",
    "一輝", "美桜", "大翔", "萌", "海斗", "七海", "蒼", "遥", "煌", "紬",
]

CLASSROOMS = [
    ("本校", 70),
    ("東校", 60),
    ("西校", 50),
]

GRADES = [
    "小4", "小5", "小6",
    "中1", "中2", "中3",
    "高1", "高2", "高3",
]

# コース別の単価マスター: (コース名, 1コマ単価, 標準教材費)
# 「単価表」シートにテーブルとして書き出し、一覧表の1コマ単価・教材費はここから
# 自動で参照する（コースが決まれば単価・教材費が自動的に決まる）。
COURSE_MASTER = [
    ("個別指導S", 4000, 7000),
    ("個別指導", 3500, 6000),
    ("集団授業", 2500, 4500),
    ("映像授業", 2100, 3000),
]
COURSE_NAMES = [c[0] for c in COURSE_MASTER]
COURSE_PRICE = {name: (unit, material) for name, unit, material in COURSE_MASTER}


def make_unique_names(count):
    """重複しない生徒名（姓+名）を count 件生成する。"""
    names = set()
    while len(names) < count:
        name = random.choice(SEI_LIST) + random.choice(MEI_LIST)
        names.add(name)
    # set→listの順序はPythonの文字列ハッシュのランダム化により実行ごとに変わる
    # (random.seedでは制御できない)ため、まずsortedで固定順にしてからshuffleする。
    names = sorted(names)
    random.shuffle(names)
    return names


def build_students():
    total = sum(n for _, n in CLASSROOMS)
    all_names = make_unique_names(total)

    students = []
    name_idx = 0
    for classroom, count in CLASSROOMS:
        for _ in range(count):
            name = all_names[name_idx]
            name_idx += 1

            grade = random.choice(GRADES)
            course = random.choice(COURSE_NAMES)
            # 週あたりの通塾回数（実際の登録データに相当する、人が入力する値）
            weekly_lessons = random.randint(1, 5)

            students.append(
                {
                    "name": name,
                    "classroom": classroom,
                    "grade": grade,
                    "course": course,
                    "weekly_lessons": weekly_lessons,
                    "email": f"seito{len(students) + 1:03d}@example.com",
                }
            )
    return students


# ------------------------------------------------------------
# Excel 出力
# ------------------------------------------------------------

HEADERS = [
    "No", "生徒名", "教室", "学年", "コース",
    "1コマ単価", "週コマ数", "コマ数", "教材費", "授業料", "請求額(税込)", "保護者メール",
]

MONEY_FORMAT = "#,##0"
TABLE_NAME = "SeikyuData"
TANKA_TABLE_NAME = "TankaData"
JUKU_NAME = "さくら学習塾"
SITE_URL = "https://ozakiray0619-000.github.io/juku-seikyu-ichiranhyo/"

# 対象年月（このスクリプトを実行した月分の請求として生成する。ダミーデータなので
# 固定文言にしたい場合はここを直接書き換えてもよい）。
_TODAY = datetime.date.today()
BILLING_PERIOD = f"{_TODAY.year}年{_TODAY.month}月分"

# 振込先・支払期日（ダミーデータ。実運用では実際の口座情報に差し替える）。
BANK_INFO = "〇〇銀行 〇〇支店　普通　1234567　口座名義：サクラガクシュウジュク"
PAYMENT_DUE_DAYS = 14  # 発行日から何日以内に振込むか

TOTAL_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
THIN = Side(style="thin", color="FFB0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 「ここは入力してよいセル」を示す統一の黄色（他の自動計算セルとは触ってよいかどうかで区別する）
INPUT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
INPUT_BORDER = Border(
    left=Side(style="medium", color="FF223757"),
    right=Side(style="medium", color="FF223757"),
    top=Side(style="medium", color="FF223757"),
    bottom=Side(style="medium", color="FF223757"),
)


def add_note(ws, cell_ref, text, width=220, height=70):
    """初心者向けの操作ヒントをセルコメントとして添える。"""
    comment = Comment(text, "使い方ガイド")
    comment.width = width
    comment.height = height
    ws[cell_ref].comment = comment


def build_gmail_url_formula(
    email_expr, name_expr, classroom_expr, course_expr,
    lessons_expr, tuition_expr, material_expr, invoice_expr,
):
    """宛先・件名・本文が入力済みのGmail作成画面を開くURLの数式を組み立てる。

    一覧表シート（行参照）・請求書シート（構造化参照の検索結果）の
    どちらからも同じ文面を組み立てられるよう、各値をセル参照の文字列として受け取る。
    """
    # 件名は短くしておく（塾名まで入れるとURLエンコード後に長くなり、
    # HYPERLINK関数の255文字制限を超えて#VALUE!エラーになりやすいため）。
    # 塾名は本文の書き出しに入っているので、件名では省略しても問題ない。
    subject_formula = f'{name_expr} & "様 {BILLING_PERIOD}請求のお知らせ"'
    tax_formula = f'ROUND(({tuition_expr}+{material_expr})*0.1,0)'
    body_formula = (
        f'{name_expr}&"様"&CHAR(10)&CHAR(10)&'
        f'"いつもお世話になっております。{JUKU_NAME}です。"&CHAR(10)&'
        f'"{BILLING_PERIOD}の請求内容は以下の通りです。"&CHAR(10)&CHAR(10)&'
        f'"教室：　"&{classroom_expr}&CHAR(10)&'
        f'"コース：　"&{course_expr}&CHAR(10)&'
        f'"コマ数：　"&{lessons_expr}&"コマ"&CHAR(10)&'
        f'"授業料：　"&TEXT({tuition_expr},"#,##0")&"円"&CHAR(10)&'
        f'"教材費：　"&TEXT({material_expr},"#,##0")&"円"&CHAR(10)&'
        f'"消費税：　"&TEXT({tax_formula},"#,##0")&"円"&CHAR(10)&'
        f'"----------------"&CHAR(10)&'
        f'"ご請求額：　"&TEXT({invoice_expr},"#,##0")&"円"&CHAR(10)&CHAR(10)&'
        f'"振込先：　{BANK_INFO}"&CHAR(10)&'
        f'"支払期日：　"&TEXT(TODAY()+{PAYMENT_DUE_DAYS},"yyyy/mm/dd")&CHAR(10)&CHAR(10)&'
        f'"お手数ですが、期日までにお振込みくださいますようお願いいたします。"'
    )
    # 本文まで含めるとURLエンコード後に長くなりすぎ、Excelの HYPERLINK関数の
    # リンク先文字数制限（255文字）を超えて #VALUE! エラーになるため、
    # リンクには宛先・件名だけを入れる（本文は別セルのプレビューから手動でコピペする）。
    gmail_url_formula = (
        f'"https://mail.google.com/mail/?view=cm&fs=1&to=" & '
        f'_xlfn.ENCODEURL({email_expr}) & "&su=" & _xlfn.ENCODEURL({subject_formula})'
    )
    return subject_formula, body_formula, gmail_url_formula


def build_gmail_url_static(student):
    """一覧表シートの各行用に、宛先・件名・請求内容まで入力済みのGmail URLを
    Python側であらかじめ組み立てる（文章ではなく項目と数字だけの請求内容）。

    Excelの HYPERLINK関数には文字数制限があり、日本語を含む本文まで
    数式で組み立てるとその制限を超えてしまう。ここではファイル生成時点の
    金額でURLを確定させることでその制限を回避する（＝週コマ数などを
    あとからExcel上で変更しても、このURLの中身は自動更新されない。
    最新の内容で送りたい場合はスクリプトを再実行してファイルを作り直す）。
    """
    unit_price, material_fee = COURSE_PRICE[student["course"]]
    lessons = student["weekly_lessons"] * 4
    tuition = unit_price * lessons
    tax = round((tuition + material_fee) * 0.1)
    invoice = tuition + material_fee + tax

    subject = f'{student["name"]}様 {BILLING_PERIOD}請求のお知らせ'
    body = (
        f'教室：{student["classroom"]}／'
        f'コース：{student["course"]}／'
        f'コマ数：{lessons}コマ／'
        f'授業料：{tuition:,}円／'
        f'教材費：{material_fee:,}円／'
        f'消費税：{tax:,}円／'
        f'請求額：{invoice:,}円'
    )
    return (
        "https://mail.google.com/mail/?view=cm&fs=1"
        f'&to={quote(student["email"])}'
        f"&su={quote(subject)}"
        f"&body={quote(body)}"
    )


def build_workbook(students):
    wb = Workbook()
    # openpyxl は数式の計算結果をキャッシュしないため、開いた瞬間に必ず
    # 再計算されるようにする(古い/空の値のまま表示されるのを防ぐ)。
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    ws = wb.active
    ws.title = "月末請求一覧"

    n_students = len(students)
    header_row = 3
    data_start_row = header_row + 1
    data_end_row = data_start_row + n_students - 1
    total_row = data_end_row + 1
    n_cols = len(HEADERS)

    # --- 合計金額（表のいちばん上、構造化参照で常に最新のテーブル範囲を集計） ---
    invoice_header = "請求額(税込)"
    ws.cell(row=1, column=1, value="合計金額（請求額の合計）")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    total_cell = ws.cell(
        row=1,
        column=2,
        value=f"=SUM({TABLE_NAME}[{invoice_header}])",
    )
    total_cell.number_format = MONEY_FORMAT
    total_cell.font = Font(bold=True, size=12, color="FFC00000")
    for col in range(1, n_cols + 1):
        ws.cell(row=1, column=col).fill = SUMMARY_FILL
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 22

    # --- ガイド行（初めて開いた人向けの一行案内） ---
    guide_cell = ws.cell(
        row=2,
        column=1,
        value=(
            "🟨 黄色いセル（週コマ数）だけ入力してください。他の列は自動計算されるので触らなくてOKです。"
            "　📧 「Gmailを開く」で宛先・件名・請求内容まで入力済みのGmailを開けます"
            "（この内容はファイルを作った時点の金額で固定。週コマ数を変えたら"
            "一覧表.pyを再実行して作り直してください）。"
            "　🔽 180人分の行は普段は折りたたんであります。「合計」行の左にある［+］を"
            "クリックすると展開して週コマ数を編集できます。"
        ),
    )
    guide_cell.font = Font(size=10, italic=True, color="FF5B6472")
    guide_cell.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    ws.row_dimensions[2].height = 18

    # --- 見出し行（配色・帯縞・フィルターはテーブルスタイルが担当） ---
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Gmail送信列の見出し（テーブル範囲の外、テーブルには含めない） ---
    gmail_header = ws.cell(row=header_row, column=13, value="Gmail")
    gmail_header.alignment = Alignment(horizontal="center", vertical="center")
    gmail_header.font = Font(bold=True)
    add_note(
        ws,
        f"{get_column_letter(13)}{header_row}",
        "📧 クリックすると、この生徒の保護者宛てに\n"
        "宛先・件名・請求内容（教室／コース／コマ数／金額など）が\n"
        "入力済みの状態でGmailの作成画面が開きます。\n"
        "※このリンクの中身はファイルを作った時点の金額で固定です。\n"
        "週コマ数を変えたあとに送る場合は、一覧表.pyを再実行して\n"
        "ファイルを作り直してから開いてください。",
    )

    # --- 請求内容（ひと目で分かる自動計算テキスト。現在の最新の金額を確認する用） ---
    content_header = ws.cell(row=header_row, column=14, value="請求内容")
    content_header.alignment = Alignment(horizontal="center", vertical="center")
    content_header.font = Font(bold=True)
    add_note(
        ws,
        f"{get_column_letter(14)}{header_row}",
        "📋 この生徒の「現時点の」請求内容です（週コマ数を変えると自動更新）。\n"
        "Gmailの本文はファイル作成時点の金額で固定されているため、\n"
        "週コマ数を変えた後はここの内容と一致しているか確認してください。",
    )
    add_note(
        ws,
        f"{get_column_letter(7)}{header_row}",
        "✏ 入力してよいのはこの列だけです。\n"
        "1週間に何コマ通うかを1〜7の数字で入力してください。\n"
        "コマ数・授業料・請求額は自動で計算されます。",
    )

    # --- 請求書シートの生徒選択（No－生徒名）用の見出し（テーブル範囲の外） ---
    # 生徒名だけだと同姓同名がいた場合に選択が一意に定まらないため、
    # 一意な No を必ず含んだ形式の値をドロップダウンの選択肢にする。
    picker_header = ws.cell(row=header_row, column=15, value="選択用(No：生徒名)")
    picker_header.alignment = Alignment(horizontal="center", vertical="center")
    picker_header.font = Font(bold=True)
    add_note(
        ws,
        f"{get_column_letter(15)}{header_row}",
        "🔎 請求書シートの「②生徒名で選択」ドロップダウンの選択肢です。\n"
        "同姓同名がいても一意に選べるよう、No付きの表記にしています。\n"
        "このセル自体を直接編集する必要はありません。",
    )

    # --- データ行 ---
    # 列: A=No B=生徒名 C=教室 D=学年 E=コース F=1コマ単価 G=週コマ数
    #     H=コマ数 I=教材費 J=授業料 K=請求額(税込) L=保護者メール
    for i, student in enumerate(students):
        row = data_start_row + i
        # No は行位置から自動計算する数式にして、行の挿入/削除に追随させる
        ws.cell(row=row, column=1, value=f"=ROW()-{header_row}")
        ws.cell(row=row, column=2, value=student["name"])
        ws.cell(row=row, column=3, value=student["classroom"])
        ws.cell(row=row, column=4, value=student["grade"])
        ws.cell(row=row, column=5, value=student["course"])

        # 1コマ単価・教材費は「単価表」シートからコースに応じて自動参照する
        c_unit = ws.cell(
            row=row,
            column=6,
            value=(
                f'=IFERROR(INDEX({TANKA_TABLE_NAME}[1コマ単価],'
                f'MATCH(E{row},{TANKA_TABLE_NAME}[コース],0)),0)'
            ),
        )
        c_weekly = ws.cell(row=row, column=7, value=student["weekly_lessons"])
        # コマ数(月間)は「週コマ数×4週」で自動計算
        c_lessons = ws.cell(row=row, column=8, value=f"=G{row}*4")
        c_material = ws.cell(
            row=row,
            column=9,
            value=(
                f'=IFERROR(INDEX({TANKA_TABLE_NAME}[教材費(標準)],'
                f'MATCH(E{row},{TANKA_TABLE_NAME}[コース],0)),0)'
            ),
        )

        c_tuition = ws.cell(row=row, column=10, value=f"=F{row}*H{row}")
        c_invoice = ws.cell(
            row=row,
            column=11,
            value=f"=J{row}+I{row}+ROUND((J{row}+I{row})*0.1,0)",
        )

        ws.cell(row=row, column=12, value=student["email"])

        # 保護者へのGmail送信リンク（M列）。宛先・件名・請求内容までこの時点の
        # 金額で確定させたURLを、数式ではなく普通のハイパーリンクとして埋め込む
        # （数式のHYPERLINK関数だと本文まで含めた場合の文字数制限に引っかかるため）。
        gmail_cell = ws.cell(row=row, column=13, value="📧 Gmailを開く")
        gmail_cell.hyperlink = build_gmail_url_static(student)
        gmail_cell.font = Font(color="FF1155CC", underline="single")
        gmail_cell.alignment = Alignment(horizontal="center")

        # 請求内容サマリー（N列）。文章にはせず、項目と数字だけを並べる。
        # HYPERLINK関数の文字数制限を受けない普通のテキストセルなので、
        # 生徒名が長くても金額が大きくても安全に全文表示できる。
        content_cell = ws.cell(
            row=row,
            column=14,
            value=(
                f'="教室："&C{row}&"／コース："&E{row}&"／コマ数："&H{row}&"コマ'
                f'／授業料："&TEXT(J{row},"#,##0")&"円'
                f'／教材費："&TEXT(I{row},"#,##0")&"円'
                f'／消費税："&TEXT(ROUND((J{row}+I{row})*0.1,0),"#,##0")&"円'
                f'／請求額："&TEXT(K{row},"#,##0")&"円"'
            ),
        )
        content_cell.font = Font(size=10, color="FF3B4453")

        # 請求書シートの生徒選択ドロップダウン用（O列）。No付きなので同姓同名でも一意。
        ws.cell(row=row, column=15, value=f'=A{row}&"："&B{row}')

        for c in (c_unit, c_material, c_tuition, c_invoice):
            c.number_format = MONEY_FORMAT
        c_weekly.number_format = "0"
        c_lessons.number_format = "0"

        # 週コマ数だけが人が入力する値なので、黄色で目立たせる
        c_weekly.fill = INPUT_FILL
        c_weekly.alignment = Alignment(horizontal="center")

        for col in (1, 3, 4, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    # --- 週コマ数の入力規則（1〜7の数字以外を入れるとエラー表示） ---
    dv_weekly = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=7,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="週コマ数は1〜7の数字で入力してください。",
    )
    ws.add_data_validation(dv_weekly)
    dv_weekly.add(f"G{data_start_row}:G{data_end_row}")

    # --- 見出し+データ範囲を Excel テーブル（ListObject）化 ---
    table_ref = f"A{header_row}:{get_column_letter(n_cols)}{data_end_row}"
    table_columns = [
        TableColumn(id=idx, name=title) for idx, title in enumerate(HEADERS, start=1)
    ]
    tab = Table(displayName=TABLE_NAME, ref=table_ref, tableColumns=table_columns)
    tab.autoFilter = AutoFilter(ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # --- 合計行（表のいちばん下、構造化参照で行の増減に自動追随） ---
    ws.cell(row=total_row, column=1, value="合計")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

    for header in ("コマ数", "教材費", "授業料", "請求額(税込)"):
        col = HEADERS.index(header) + 1
        cell = ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({TABLE_NAME}[{header}])",
        )
        cell.number_format = MONEY_FORMAT
        cell.font = Font(bold=True)

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = BORDER

    # --- 180行のデータ行をグループ化して折りたたむ（Excelのアウトライン機能） ---
    # タブを開いたときに180行がずらっと表示されると見づらいので、普段は折りたたんで
    # おき、「合計」行の左の［+］をクリックしたときだけ展開して週コマ数を編集できる
    # ようにする。以前は非表示にしていたが、それだと入力欄ごと迷子になったため、
    # 「折りたたみ」なら見た目はすっきりしたまま、入力できる場所も分かりやすい。
    ws.sheet_properties.outlinePr = Outline(summaryBelow=True, showOutlineSymbols=True)
    for row in range(data_start_row, data_end_row + 1):
        ws.row_dimensions[row].outlineLevel = 1
        ws.row_dimensions[row].hidden = True
    ws.row_dimensions[total_row].collapsed = True

    # --- 列幅 ---
    widths = [6, 14, 8, 6, 12, 11, 9, 8, 10, 12, 14, 22, 16, 62, 20]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- 見出し行を固定表示 ---
    ws.freeze_panes = f"A{data_start_row}"

    invoice_ws = build_invoice_sheet(wb, ws.title, data_start_row, data_end_row)

    # 月末請求一覧シートは請求書シート・単価表シートの計算元データであると同時に、
    # このツールで人が入力する唯一の値（週コマ数、黄色いセル）を触れる唯一の場所
    # でもあるため、非表示にはしない（以前は180行のスクロールを嫌って非表示にして
    # いたが、そうすると週コマ数の入力欄ごと迷子になってしまった）。代わりに180行は
    # アウトライン機能で折りたたんでおき、開いたときの見た目はすっきりさせつつ、
    # ワークブックを開いたときに最初に表示されるのは請求書シートにする。
    wb.active = wb.index(invoice_ws)

    return wb


def build_tanka_sheet(ws, start_row, start_col):
    """コース別の単価マスター表を、指定シートの指定位置に埋め込む。

    請求書シートの操作パネル内（印刷範囲の外）に置くことで、単価表を
    別タブに分けずに1シートで完結させる。ここを直せば一覧表の該当コース
    全員に自動反映される（TANKA_TABLE_NAME はシートをまたいで参照される
    ため、物理的な置き場所はどこでも構わない）。
    """
    title_row = start_row
    ws.cell(row=title_row, column=start_col, value="📋 コース別 単価マスター").font = Font(
        bold=True, size=12, color="FF223757"
    )
    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=start_col + 2)

    note_row = title_row + 1
    ws.cell(
        row=note_row,
        column=start_col,
        value="🟨 黄色いセル（1コマ単価・教材費）を直すと、一覧表の該当コースの生徒全員に自動反映されます。",
    ).font = Font(size=9, color="FF5B6472")
    ws.cell(row=note_row, column=start_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=start_col, end_row=note_row + 1, end_column=start_col + 2)

    tanka_headers = ["コース", "1コマ単価", "教材費(標準)"]
    header_row = note_row + 2
    for col_offset, title in enumerate(tanka_headers):
        ws.cell(row=header_row, column=start_col + col_offset, value=title)
    add_note(
        ws,
        f"{get_column_letter(start_col + 1)}{header_row}",
        "✏ ここが入力欄です。\n"
        "数字を書き換えると、一覧表の該当コースを選んでいる生徒全員の金額に自動で反映されます。",
    )

    for i, (course, unit_price, material_fee) in enumerate(COURSE_MASTER):
        row = header_row + 1 + i
        ws.cell(row=row, column=start_col, value=course)
        c_unit = ws.cell(row=row, column=start_col + 1, value=unit_price)
        c_material = ws.cell(row=row, column=start_col + 2, value=material_fee)
        c_unit.number_format = MONEY_FORMAT
        c_material.number_format = MONEY_FORMAT
        c_unit.fill = INPUT_FILL
        c_material.fill = INPUT_FILL

    data_end_row = header_row + len(COURSE_MASTER)
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(start_col + 2)
    table_ref = f"{start_col_letter}{header_row}:{end_col_letter}{data_end_row}"
    table_columns = [
        TableColumn(id=idx, name=title) for idx, title in enumerate(tanka_headers, start=1)
    ]
    tab = Table(displayName=TANKA_TABLE_NAME, ref=table_ref, tableColumns=table_columns)
    tab.autoFilter = AutoFilter(ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    return data_end_row


def build_invoice_sheet(wb, main_sheet_name, data_start_row, data_end_row):
    """生徒名を選ぶとその生徒の個人請求書とメール送信用リンクが自動生成されるシート。

    列A〜Fが「そのまま渡せる正式な請求書」（印刷範囲もここだけに絞る）、
    列H以降が「Noや生徒名を入力する操作パネル」（印刷・PDFには含まれない）。
    左右を分けることで、保護者に渡すPDFに操作用の入力欄やGmailリンクが
    写り込まないようにしている。
    """
    ws = wb.create_sheet("請求書")
    ws.sheet_view.showGridLines = False

    PANEL_LABEL_COL = 8   # H列
    PANEL_VALUE_COL = 10  # J列（J:Lを結合して使う）

    # データ入力規則（ドロップダウン・数値範囲）は、別シートの「テーブルの構造化参照」
    # （例: SeikyuData[生徒名]）を直接読み込めない（Excelの既知の制限）。
    # 必ずプレーンなセル範囲参照（例: 月末請求一覧!$B$4:$B$183）を使う。
    # ②のドロップダウンは生徒名そのものではなく「No：生徒名」列（O列）を参照する。
    # 生徒名だけを選択肢にすると同姓同名がいた場合にどちらを選んだか区別できず、
    # 別人の請求内容が表示されてしまうため、必ず一意な No を経由して生徒を特定する。
    name_range_ref = f"'{main_sheet_name}'!$O${data_start_row}:$O${data_end_row}"

    NO_CELL = f"{get_column_letter(PANEL_VALUE_COL)}4"
    NO_CELL_ABS = f"${get_column_letter(PANEL_VALUE_COL)}$4"
    NAME_INPUT_CELL = f"{get_column_letter(PANEL_VALUE_COL)}5"
    NAME_INPUT_ABS = f"${get_column_letter(PANEL_VALUE_COL)}$5"
    RESOLVED_NO_ABS = f"${get_column_letter(PANEL_VALUE_COL)}$6"    # ①②どちらの入力からでも解決される選択中の No（数値）
    RESOLVED_NAME_ABS = f"${get_column_letter(PANEL_VALUE_COL)}$7"  # 上のNoに対応する生徒名（表示用）

    def lookup(header):
        # 生徒名ではなく一意な No で引く（同姓同名がいても正しい行を取得できるように）。
        return f'IFERROR(INDEX({TABLE_NAME}[{header}],MATCH({RESOLVED_NO_ABS},{TABLE_NAME}[No],0)),"")'

    # ============================================================
    # 左側：正式な請求書（A〜F列。印刷範囲もここだけに絞る）
    # ============================================================
    title = ws.cell(row=1, column=1, value="請求書")
    title.font = Font(bold=True, size=26, color="FF223757")
    ws.merge_cells("A1:C1")

    ws.cell(row=1, column=4, value="発行日：").alignment = Alignment(horizontal="right")
    issue_date = ws.cell(row=1, column=5, value="=TODAY()")
    issue_date.number_format = "yyyy/mm/dd"
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)

    ws.cell(
        row=2, column=1, value=f"{JUKU_NAME}　{BILLING_PERIOD}請求書"
    ).font = Font(size=11, color="FF5B6472")

    ws.cell(row=2, column=4, value="請求書No：").alignment = Alignment(horizontal="right")
    invoice_no_cell = ws.cell(row=2, column=5, value=f"={lookup('No')}")
    invoice_no_cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)

    recipient_cell = ws.cell(
        row=4,
        column=1,
        value=(
            f'=IF({RESOLVED_NO_ABS}="","（右の操作パネルでNoか生徒名を選択してください）",'
            f'{RESOLVED_NAME_ABS}&"　様")'
        ),
    )
    recipient_cell.font = Font(bold=True, size=16, color="FF223757")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

    ws.cell(row=5, column=1, value="下記の通りご請求申し上げます。").font = Font(size=10)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)

    # --- 明細 ---
    detail_rows = [
        ("教室", lookup("教室"), None),
        ("学年", lookup("学年"), None),
        ("コース", lookup("コース"), None),
        ("1コマ単価", lookup("1コマ単価"), MONEY_FORMAT),
        ("週コマ数", lookup("週コマ数"), "0"),
        ("コマ数（月間）", lookup("コマ数"), "0"),
        ("教材費", lookup("教材費"), MONEY_FORMAT),
        ("授業料", lookup("授業料"), MONEY_FORMAT),
        (
            "消費税（授業料＋教材費の10%）",
            f'IFERROR(ROUND((INDEX({TABLE_NAME}[授業料],MATCH({RESOLVED_NO_ABS},{TABLE_NAME}[No],0))+'
            f'INDEX({TABLE_NAME}[教材費],MATCH({RESOLVED_NO_ABS},{TABLE_NAME}[No],0)))*0.1,0),"")',
            MONEY_FORMAT,
        ),
        ("ご請求額（税込）", lookup("請求額(税込)"), MONEY_FORMAT),
    ]

    start_row = 8
    ws.cell(row=start_row - 1, column=1, value="ご請求内容").font = Font(bold=True, size=12)
    for i, (label, formula, number_format) in enumerate(detail_rows):
        r = start_row + i
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = Font(bold=(label == "ご請求額（税込）"))
        value_cell = ws.cell(row=r, column=3, value=f"={formula}")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        if number_format:
            value_cell.number_format = number_format
        if label == "ご請求額（税込）":
            value_cell.font = Font(bold=True, size=14, color="FFC00000")
            label_cell.font = Font(bold=True, size=12)
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = PatternFill(
                    start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
                )

    total_row_ref = start_row + len(detail_rows) - 1  # ご請求額の行

    footer_row = total_row_ref + 2
    ws.cell(row=footer_row, column=1, value="振込先：").font = Font(size=10, bold=True)
    bank_cell = ws.cell(row=footer_row, column=2, value=BANK_INFO)
    bank_cell.font = Font(size=10)
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=6)

    due_row = footer_row + 1
    ws.cell(row=due_row, column=1, value="支払期日：").font = Font(size=10, bold=True)
    due_cell = ws.cell(row=due_row, column=2, value=f"=TODAY()+{PAYMENT_DUE_DAYS}")
    due_cell.number_format = "yyyy/mm/dd"
    due_cell.font = Font(size=10)
    ws.merge_cells(start_row=due_row, start_column=2, end_row=due_row, end_column=6)

    thanks_row = due_row + 1
    ws.cell(
        row=thanks_row,
        column=1,
        value="お手数ですが、期日までにお振込みくださいますようお願いいたします。",
    ).font = Font(size=10)
    ws.merge_cells(start_row=thanks_row, start_column=1, end_row=thanks_row, end_column=6)

    close_row = thanks_row + 2
    ws.cell(row=close_row, column=1, value="以上").font = Font(size=10)
    ws.cell(row=close_row, column=4, value="発行者：").alignment = Alignment(horizontal="right")
    ws.cell(row=close_row, column=5, value=JUKU_NAME)
    ws.merge_cells(start_row=close_row, start_column=5, end_row=close_row, end_column=6)

    # --- 印刷範囲（この請求書をそのままPDFにできるように） ---
    # 「ファイル」→「エクスポート」→「PDF/XPSドキュメントの作成」、
    # または Ctrl+P →「PDFとして保存」で、A〜F列・この範囲だけが1枚のPDFになる。
    # 右側（H列以降）の操作パネルは印刷範囲の外なので、書き出したPDFには写らない。
    invoice_last_row = close_row + 1
    ws.print_area = f"A1:F{invoice_last_row}"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "portrait"

    # ============================================================
    # 右側：操作パネル（H列以降。印刷・PDFには含まれない）
    # ============================================================
    PL, PV = PANEL_LABEL_COL, PANEL_VALUE_COL
    panel_header = ws.cell(
        row=1, column=PL, value="🔧 操作パネル（この部分はPDF・印刷には含まれません）"
    )
    panel_header.font = Font(bold=True, size=10, color="FF5B6472")
    panel_header.fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    ws.merge_cells(start_row=1, start_column=PL, end_row=1, end_column=PV + 2)

    ws.cell(
        row=2, column=PL, value="🟨 黄色いセルが入力欄です"
    ).font = Font(size=9, italic=True, color="FF5B6472")

    # クリックの動作確認用に、生徒選択に関係なく常に表示される固定のGmailリンクを用意する
    # (openpyxl のネイティブハイパーリンク機能。数式のHYPERLINK関数とは別の仕組みで、
    # こちらが確実に動くかどうかで問題の切り分けができる)
    test_link_cell = ws.cell(row=3, column=PL, value="✅ クリックテスト：Gmailを開く（動作確認用）")
    test_link_cell.hyperlink = "https://mail.google.com/mail/?view=cm&fs=1&su=クリックテスト"
    test_link_cell.font = Font(bold=True, color="FF1155CC", underline="single")
    ws.merge_cells(start_row=3, start_column=PL, end_row=3, end_column=PV + 1)

    # --- 全生徒を検索できるWebページへのリンク（月末請求一覧シートを開かなくても
    # ブラウザでパッと検索・絞り込みしたいときはこちらから） ---
    SITE_COL = PV + 4  # N列
    site_header = ws.cell(row=1, column=SITE_COL, value="🔍 生徒を探すときは")
    site_header.font = Font(bold=True, size=10, color="FF5B6472")
    ws.merge_cells(start_row=1, start_column=SITE_COL, end_row=1, end_column=SITE_COL + 2)

    site_link_cell = ws.cell(row=2, column=SITE_COL, value="🔗 一覧をWebで検索して見る")
    site_link_cell.hyperlink = SITE_URL
    site_link_cell.font = Font(bold=True, size=12, color="FF1155CC", underline="single")
    site_link_cell.fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    ws.merge_cells(start_row=2, start_column=SITE_COL, end_row=2, end_column=SITE_COL + 2)

    site_note_cell = ws.cell(
        row=3,
        column=SITE_COL,
        value="生徒名やNoで検索、教室で絞り込みができます（ブラウザで開きます）。",
    )
    site_note_cell.font = Font(size=9, color="FF5B6472")
    site_note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=SITE_COL, end_row=4, end_column=SITE_COL + 2)

    # --- ① 生徒番号(No)で入力 ---
    ws.cell(row=4, column=PL, value="① 生徒番号(No)で入力 ▶").font = Font(bold=True)
    no_cell = ws.cell(row=4, column=PV, value=None)
    no_cell.font = Font(bold=True, size=14)
    no_cell.fill = INPUT_FILL
    no_cell.border = INPUT_BORDER
    no_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=4, start_column=PV, end_row=4, end_column=PV + 1)
    add_note(
        ws,
        NO_CELL,
        "✏ ここに一覧表シートの「No」（生徒番号）を入力してください。\n"
        "①か②のどちらか一方に入力すればOKです。",
    )

    dv_no = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=f"=COUNTA({name_range_ref})",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="一覧表にあるNo（生徒番号）を入力してください。",
    )
    ws.add_data_validation(dv_no)
    dv_no.add(NO_CELL)

    # --- ② または 生徒名で選択 ---
    ws.cell(row=5, column=PL, value="② または生徒名で選択 ▶").font = Font(bold=True)
    name_cell = ws.cell(row=5, column=PV, value=None)
    name_cell.font = Font(bold=True, size=14)
    name_cell.fill = INPUT_FILL
    name_cell.border = INPUT_BORDER
    ws.merge_cells(start_row=5, start_column=PV, end_row=5, end_column=PV + 2)
    add_note(
        ws,
        NAME_INPUT_CELL,
        "✏ セルをクリックすると▼が出るので、「No：生徒名」の形式から選んでください。\n"
        "同姓同名がいてもNo付きなので取り違えずに選べます。",
    )

    dv_name = DataValidation(
        type="list",
        formula1=f"={name_range_ref}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="一覧表にある「No：生徒名」から選択してください。",
    )
    ws.add_data_validation(dv_name)
    dv_name.add(NAME_INPUT_CELL)

    # --- 選択中の生徒（①②どちらの入力からでも解決） ---
    # No（数値）で解決してから生徒名を引く。生徒名を直接キーにすると同姓同名の
    # 場合に別人の行にマッチしてしまうため、必ずNoを経由する。
    ws.cell(row=6, column=PL, value="選択中のNo").font = Font(bold=True)
    resolved_no_cell = ws.cell(
        row=6,
        column=PV,
        value=(
            f'=IF({NO_CELL_ABS}<>"",'
            f'IF(COUNTIF({TABLE_NAME}[No],{NO_CELL_ABS})=0,"⚠ 該当するNoがありません",{NO_CELL_ABS}),'
            f'IF({NAME_INPUT_ABS}<>"",'
            f'VALUE(LEFT({NAME_INPUT_ABS},FIND("：",{NAME_INPUT_ABS})-1)),""))'
        ),
    )
    resolved_no_cell.font = Font(bold=True, size=13, color="FF223757")
    ws.merge_cells(start_row=6, start_column=PV, end_row=6, end_column=PV + 2)

    ws.cell(row=7, column=PL, value="選択中の生徒名").font = Font(bold=True)
    resolved_name_cell = ws.cell(row=7, column=PV, value=f"={lookup('生徒名')}")
    resolved_name_cell.font = Font(bold=True, size=13, color="FF223757")
    ws.merge_cells(start_row=7, start_column=PV, end_row=7, end_column=PV + 2)

    # --- メール本文（件名・本文の元になる文面。件名行にも本文コピー行にも使い回す） ---
    subject_formula, body_core_formula, _ = build_gmail_url_formula(
        email_expr=lookup("保護者メール"),
        name_expr=RESOLVED_NAME_ABS,
        classroom_expr=lookup("教室"),
        course_expr=lookup("コース"),
        lessons_expr=lookup("コマ数"),
        tuition_expr=lookup("授業料"),
        material_expr=lookup("教材費"),
        invoice_expr=lookup("請求額(税込)"),
    )
    body_formula = (
        f'IF({RESOLVED_NO_ABS}="","生徒名を選択すると本文が生成されます。",{body_core_formula})'
    )

    # --- メール送信 ---
    email_row = 9
    ws.cell(row=email_row - 1, column=PL, value="メールで送る").font = Font(bold=True, size=12)

    ws.cell(row=email_row, column=PL, value="宛先（保護者メール）")
    ws.cell(row=email_row, column=PV, value=f"={lookup('保護者メール')}")
    ws.merge_cells(start_row=email_row, start_column=PV, end_row=email_row, end_column=PV + 2)

    subject_row = email_row + 1
    ws.cell(row=subject_row, column=PL, value="件名")
    ws.cell(row=subject_row, column=PV, value=f"={subject_formula}")
    ws.merge_cells(start_row=subject_row, start_column=PV, end_row=subject_row, end_column=PV + 2)

    # Gmail の作成画面を直接開く専用URL（mailto: と違い、OS/ブラウザ側の
    # メールアプリ設定が一切不要 — 普通の https リンクとしてブラウザで開く）。
    # 本文まで含めるとURLが長くなりすぎてHYPERLINK関数の255文字制限で
    # #VALUE! エラーになるため、宛先・件名だけをリンクに入れる
    # （本文は下の「本文プレビュー」からコピペする）。
    gmail_url_formula = (
        f'"https://mail.google.com/mail/?view=cm&fs=1&to=" & '
        f'_xlfn.ENCODEURL({lookup("保護者メール")}) & "&su=" & _xlfn.ENCODEURL({subject_formula})'
    )
    gmail_link_formula = (
        f'IF({RESOLVED_NO_ABS}="","",'
        f'HYPERLINK({gmail_url_formula},"📧 Gmail作成画面を開く（宛先・件名 入力済み）"))'
    )
    link_row = subject_row + 1
    link_cell = ws.cell(row=link_row, column=PV, value=f"={gmail_link_formula}")
    link_cell.font = Font(bold=True, size=12, color="FF1155CC", underline="single")
    link_cell.fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    ws.merge_cells(start_row=link_row, start_column=PV, end_row=link_row, end_column=PV + 2)

    note_row = link_row + 1
    ws.cell(
        row=note_row,
        column=PV,
        value=(
            "※このリンクは宛先・件名だけが入力された状態でGmailを開きます"
            "（本文まで入れるとURLが長くなりすぎてエラーになるため）。"
            "開いたら、下の「本文プレビュー」をコピーして本文欄に貼り付けてください。"
            "※Googleにログイン済みの状態でブラウザが開く必要があります。"
        ),
    ).font = Font(size=9, color="FF5B6472")
    ws.cell(row=note_row, column=PV).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=PV, end_row=note_row, end_column=PV + 2)

    # --- 本文プレビュー（コピーもできる） ---
    body_row = note_row + 2
    ws.cell(row=body_row, column=PL, value="本文プレビュー（コピーも可）").font = Font(bold=True)

    # 本文は15行前後（挨拶＋明細＋結びの文）あるため、8行分だと文末が
    # はみ出して見えなくなる。余裕を持って16行分確保する。
    body_cell = ws.cell(row=body_row + 1, column=PL, value=f"={body_formula}")
    ws.merge_cells(start_row=body_row + 1, start_column=PL, end_row=body_row + 16, end_column=PV + 2)
    body_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    body_cell.border = BORDER

    # --- 単価表（別タブに分けず、この操作パネルの下に埋め込む） ---
    tanka_start_row = body_row + 18
    build_tanka_sheet(ws, tanka_start_row, PV)

    # --- 列幅 ---
    widths = {
        "A": 27, "B": 3, "C": 14, "D": 14, "E": 14, "F": 14, "G": 4,
        "H": 22, "I": 2, "J": 16, "K": 16, "L": 16, "M": 3, "N": 20, "O": 12, "P": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    return ws


def compute_invoice_rows(students):
    """一覧表.xlsxの各行と同じ計算ロジックで、Web用の集計データを作る。

    xlsx側の数式（=授業料+教材費+ROUND((授業料+教材費)*0.1,0)）と必ず一致させる
    ため、単価・税計算のロジックはここでも同じ式を使う。
    """
    rows = []
    sum_by_classroom = {}
    total_invoice = 0
    for i, student in enumerate(students, start=1):
        unit_price, material_fee = COURSE_PRICE[student["course"]]
        lessons = student["weekly_lessons"] * 4
        tuition = unit_price * lessons
        tax = round((tuition + material_fee) * 0.1)
        invoice = tuition + material_fee + tax
        rows.append(
            [
                i, student["name"], student["classroom"], student["grade"], student["course"],
                unit_price, student["weekly_lessons"], lessons, material_fee, tuition, invoice,
            ]
        )
        sum_by_classroom[student["classroom"]] = (
            sum_by_classroom.get(student["classroom"], 0) + invoice
        )
        total_invoice += invoice
    return rows, sum_by_classroom, total_invoice


def update_site_data(students, html_path="index.html"):
    """index.html に埋め込まれたプレビュー用データ(DATA)を、一覧表.xlsxと同じ
    生成元データ・同じ計算ロジックから作り直す。

    以前は手作業でこのDATAを更新しており、xlsxとサイトの内容がずれるバグが
    実際に起きたため、一覧表.py実行のたびに必ず自動で同期する。
    """
    rows, sum_by_classroom, total_invoice = compute_invoice_rows(students)
    data_json = json.dumps(
        {"rows": rows, "sum_by_classroom": sum_by_classroom, "total_invoice": total_invoice},
        ensure_ascii=False,
    )

    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    new_html, n = re.subn(
        r"const DATA = \{.*?\};",
        lambda _match: f"const DATA = {data_json};",
        html,
        count=1,
        flags=re.DOTALL,
    )
    if n == 0:
        raise RuntimeError(f"{html_path} 内に `const DATA = ...;` が見つかりませんでした。")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(new_html)


def main():
    students = build_students()
    wb = build_workbook(students)
    output_path = "一覧表.xlsx"
    wb.save(output_path)
    print(f"生徒{len(students)}名分の月末請求一覧表を作成しました: {output_path}")

    update_site_data(students)
    print("index.html の埋め込みデータを一覧表.xlsxと同じ内容に同期しました。")


if __name__ == "__main__":
    main()
